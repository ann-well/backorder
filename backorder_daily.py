import openpyxl
from openpyxl import Workbook
import datetime
import tkinter as tk
from tkinter import messagebox


# Create window for GUI
window = tk.Tk()
window.title('Backorder sorting')
info = tk.Text(fg="white", bg="black", width=70, height=10)
info.configure(font=('Times New Roman', 12))
info.insert('end', 'Required files:\n')
info.insert('end', 'Backorders in E1.xlsx\n')
info.insert('end', 'Consumer data for backorders.xlsx\n\n')
info.pack()

# Set column dimensions
def set_columns(excel_sheet):
    excel_sheet.column_dimensions['A'].width = 16
    excel_sheet.column_dimensions['B'].width = 13
    excel_sheet.column_dimensions['C'].width = 13
    excel_sheet.column_dimensions['D'].width = 20
    excel_sheet.column_dimensions['E'].width = 30
    excel_sheet.column_dimensions['F'].width = 13
    excel_sheet.column_dimensions['G'].width = 18
    excel_sheet.column_dimensions['H'].width = 40
    excel_sheet.column_dimensions['I'].width = 18
    excel_sheet.column_dimensions['J'].width = 18

# Open daily report file
try:
    wb = openpyxl.load_workbook('Backorders in E1.xlsx')
    sheet = wb['Page1_1']

except:  # No file error handling
    messagebox.showwarning("No file detected",
                           "No 'Backorders in E1.xlsx' found.\nPlease exit and put the file in this folder")
            
# Open cognos file
try:
    cogn = openpyxl.load_workbook('Consumer data for backorders.xlsx')
    sheet1 = cogn['Page1_1']

except:  # No file error handling
    messagebox.showwarning("No file detected",
                           "No 'Consumer data for backorders.xlsx' found.\nPlease exit and put the file in this folder")


def sortBackorder():
    # Create lists to check if order number is present more than once
    single = []
    duplicate = []

    # Iterate through list
    for i in range(3, sheet.max_row):
        garment_list = ['belt', 'boxer', 'brief', 'waistband', 'comfizz', 'fulcionel', 'underwear', 'coresitwell', 'hernia', 'level', 'vest', 'vanilla blush', 'supportx',
                        'prolapse'] # list of garment keywords

        # Skip if garment or 2180102006, or 10GB3S                                  (--ADJUST codes if needed--)
        if [garm for garm in garment_list if(garm in sheet.cell(row=i, column=7).value.lower())] or sheet.cell(row=i, column=6).value == '380P10GB3S               ' or sheet.cell(row=i, column=6).value == '2180102006':
            sheet.cell(row=i, column=1).value = 'not to do'
            continue
        if sheet.cell(row=i, column=1).value in single:
            duplicate.append(sheet.cell(row=i, column=1).value)
        else:
            single.append(sheet.cell(row=i, column=1).value)

    # Create dictionaries of email, mobile, landline and no_details
    email = {}
    mobile = {}
    landline = {}
    no_details = {}

    time_now = datetime.datetime.now()

    # Iterate through cognos report and populate dictionaries
    for i in range(2, sheet1.max_row + 1):
        # Check if patient registered less than 7 days ago - probably npr           (--ADJUST days number if needed--)
        if sheet1.cell(row=i, column=7).value is not None:
            if sheet1.cell(row=i, column=7).value - datetime.timedelta(days = 7) > time_now:
                continue 
        # Check if patient has email
        if sheet1.cell(row=i, column=2).value is not None:
            email.setdefault(sheet1.cell(row=i, column=1).value, [sheet1.cell(row=i, column=2).value, sheet1.cell(row=i, column=5).value, sheet1.cell(row=i, column=6).value, sheet1.cell(row=i, column=3).value, sheet1.cell(row=i, column=4).value])
            continue
        # Check if patient has mobile
        if sheet1.cell(row=i, column=3).value is not None:
            mobile.setdefault(sheet1.cell(row=i, column=1).value, [sheet1.cell(row=i, column=3).value.replace('+440', '44')[0:12], sheet1.cell(row=i, column=5).value, sheet1.cell(row=i, column=6).value, sheet1.cell(row=i, column=4).value])
            continue
        # Check if patient has landline
        if sheet1.cell(row=i, column=4).value is not None:
            landline.setdefault(sheet1.cell(row=i, column=1).value, [sheet1.cell(row=i, column=4).value.replace('+44', '')[0:11], sheet1.cell(row=i, column=5).value, sheet1.cell(row=i, column=6).value])
            continue
        # No contact details
        no_details.setdefault(sheet1.cell(row=i, column=1).value, [sheet1.cell(row=i, column=5).value, sheet1.cell(row=i, column=6).value])

    # Create output files
    bss_file = Workbook()
    duplicate_email_sheet = bss_file.active
    duplicate_email_sheet.title = 'Email duplicates'
    duplicate_mobile_sheet = bss_file.create_sheet(title="Mobile duplicates")
    landline_sheet = bss_file.create_sheet(title="Landline & None")
    dup_em = 1 # Iteration for emails
    dup_mob = 1 # Iteration for mobile
    land = 1 # Iteration for landline

    # Set column dimensions
    set_columns(duplicate_email_sheet)
    set_columns(duplicate_mobile_sheet)
    set_columns(landline_sheet)

    automated_file = Workbook()
    email_sheet =  automated_file.active
    email_sheet.title = "Email"
    mobile_sheet = automated_file.create_sheet(title="Mobile")
    em = 1 # Iteration for emails
    mob = 1 # Iteration for mobile

    # Set column dimensions
    set_columns(email_sheet)
    set_columns(mobile_sheet)

    # Iterate through list
    for i in range(3, sheet.max_row):
        # Garments & company
        if sheet.cell(row=i, column=1).value == 'not to do':
            continue 
        # Duplicate emails
        if sheet.cell(row=i, column=1).value in duplicate:
            if sheet.cell(row=i, column=1).value in email:
                duplicate_email_sheet.cell(row=dup_em, column=1).value = sheet.cell(row=i, column=3).value
                duplicate_email_sheet.cell(row=dup_em, column=2).value = sheet.cell(row=i, column=1).value
                duplicate_email_sheet.cell(row=dup_em, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
                duplicate_email_sheet.cell(row=dup_em, column=4).value = sheet.cell(row=i, column=6).value   
                duplicate_email_sheet.cell(row=dup_em, column=5).value = sheet.cell(row=i, column=7).value  
                duplicate_email_sheet.cell(row=dup_em, column=6).value = email[sheet.cell(row=i, column=1).value][1]
                duplicate_email_sheet.cell(row=dup_em, column=7).value = email[sheet.cell(row=i, column=1).value][2]
                duplicate_email_sheet.cell(row=dup_em, column=8).value = email[sheet.cell(row=i, column=1).value][0]
                if email[sheet.cell(row=i, column=1).value][-2] is not None:
                    duplicate_email_sheet.cell(row=dup_em, column=9).value = email[sheet.cell(row=i, column=1).value][-2].replace('+440', '44')[0:12]
                if email[sheet.cell(row=i, column=1).value][-1] is not None:
                    duplicate_email_sheet.cell(row=dup_em, column=10).value = email[sheet.cell(row=i, column=1).value][-1].replace('+44', '')[0:11]
                dup_em += 1
                continue
            # Duplicate mobile
            if sheet.cell(row=i, column=1).value in mobile:
                duplicate_mobile_sheet.cell(row=dup_mob, column=1).value = sheet.cell(row=i, column=3).value    
                duplicate_mobile_sheet.cell(row=dup_mob, column=2).value = sheet.cell(row=i, column=1).value 
                duplicate_mobile_sheet.cell(row=dup_mob, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
                duplicate_mobile_sheet.cell(row=dup_mob, column=4).value = sheet.cell(row=i, column=6).value   
                duplicate_mobile_sheet.cell(row=dup_mob, column=5).value = sheet.cell(row=i, column=7).value  
                duplicate_mobile_sheet.cell(row=dup_mob, column=6).value = mobile[sheet.cell(row=i, column=1).value][1]
                duplicate_mobile_sheet.cell(row=dup_mob, column=7).value = mobile[sheet.cell(row=i, column=1).value][2]
                duplicate_mobile_sheet.cell(row=dup_mob, column=8).value = mobile[sheet.cell(row=i, column=1).value][0]
                if mobile[sheet.cell(row=i, column=1).value][-1] is not None:
                    duplicate_mobile_sheet.cell(row=dup_mob, column=9).value = mobile[sheet.cell(row=i, column=1).value][-1].replace('+44', '')[0:11]
                dup_mob += 1
                continue

        # Landline
        if sheet.cell(row=i, column=1).value in landline:
            landline_sheet.cell(row=land, column=1).value = sheet.cell(row=i, column=3).value  
            landline_sheet.cell(row=land, column=2).value = sheet.cell(row=i, column=1).value  
            landline_sheet.cell(row=land, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
            landline_sheet.cell(row=land, column=4).value = sheet.cell(row=i, column=6).value   
            landline_sheet.cell(row=land, column=5).value = sheet.cell(row=i, column=7).value  
            landline_sheet.cell(row=land, column=6).value = landline[sheet.cell(row=i, column=1).value][1]
            landline_sheet.cell(row=land, column=7).value = landline[sheet.cell(row=i, column=1).value][2]
            landline_sheet.cell(row=land, column=8).value = landline[sheet.cell(row=i, column=1).value][0]
            land += 1
            continue
        # No details
        if sheet.cell(row=i, column=1).value in no_details:
            landline_sheet.cell(row=land, column=1).value = sheet.cell(row=i, column=3).value     
            landline_sheet.cell(row=land, column=2).value = sheet.cell(row=i, column=1).value     
            landline_sheet.cell(row=land, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
            landline_sheet.cell(row=land, column=4).value = sheet.cell(row=i, column=6).value   
            landline_sheet.cell(row=land, column=5).value = sheet.cell(row=i, column=7).value  
            landline_sheet.cell(row=land, column=6).value = no_details[sheet.cell(row=i, column=1).value][0]
            landline_sheet.cell(row=land, column=7).value = no_details[sheet.cell(row=i, column=1).value][1]
            land += 1
            continue

        # Not duplicated orders
        # Emails
        if sheet.cell(row=i, column=1).value in email:
            email_sheet.cell(row=em, column=1).value = sheet.cell(row=i, column=3).value     
            email_sheet.cell(row=em, column=2).value = sheet.cell(row=i, column=1).value     
            email_sheet.cell(row=em, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
            email_sheet.cell(row=em, column=4).value = sheet.cell(row=i, column=6).value   
            email_sheet.cell(row=em, column=5).value = sheet.cell(row=i, column=7).value  
            email_sheet.cell(row=em, column=6).value = email[sheet.cell(row=i, column=1).value][1]
            email_sheet.cell(row=em, column=7).value = email[sheet.cell(row=i, column=1).value][2]
            email_sheet.cell(row=em, column=8).value = email[sheet.cell(row=i, column=1).value][0]
            if email[sheet.cell(row=i, column=1).value][-2] is not None:
                email_sheet.cell(row=em, column=9).value = email[sheet.cell(row=i, column=1).value][-2].replace('+440', '44')[0:12]
            if email[sheet.cell(row=i, column=1).value][-1] is not None:
                email_sheet.cell(row=em, column=10).value = email[sheet.cell(row=i, column=1).value][-1].replace('+44', '')[0:11]
            em += 1
            continue
        # Mobile
        if sheet.cell(row=i, column=1).value in mobile:
            mobile_sheet.cell(row=mob, column=1).value = sheet.cell(row=i, column=3).value     
            mobile_sheet.cell(row=mob, column=2).value = sheet.cell(row=i, column=1).value     
            mobile_sheet.cell(row=mob, column=3).value = sheet.cell(row=i, column=2).value.strftime('%d-%m-%Y')
            mobile_sheet.cell(row=mob, column=4).value = sheet.cell(row=i, column=6).value   
            mobile_sheet.cell(row=mob, column=5).value = sheet.cell(row=i, column=7).value  
            mobile_sheet.cell(row=mob, column=6).value = mobile[sheet.cell(row=i, column=1).value][1]
            mobile_sheet.cell(row=mob, column=7).value = mobile[sheet.cell(row=i, column=1).value][2]
            mobile_sheet.cell(row=mob, column=8).value = mobile[sheet.cell(row=i, column=1).value][0]
            if mobile[sheet.cell(row=i, column=1).value][-1] is not None:
                mobile_sheet.cell(row=mob, column=9).value = mobile[sheet.cell(row=i, column=1).value][-1].replace('+44', '')[0:11]
            mob += 1

    # Save output files
    bss_file.save('BSS Backorder ' + time_now.strftime('%d%m%Y') + '.xlsx')
    automated_file.save('Automated Backorder ' + time_now.strftime('%d%m%Y') + '.xlsx')

    # Close window after program ends
    window.destroy()

# Start the function by user input
addButton = tk.Button(window,
                      text="START",
                      command=sortBackorder)
addButton.pack()

window.mainloop()

